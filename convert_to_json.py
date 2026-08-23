import json
import re
from datetime import datetime, timezone
from openpyxl import load_workbook

SRC = "/mnt/user-data/uploads/Master_Kecamatan_Tervalidasi_v7_Final.xlsx"
OUT = "data_source/data_koordinat.json"

# Provinsi -> Zona Waktu Indonesia
WIB = {
    "Aceh", "Sumatera Utara", "Sumatera Barat", "Riau", "Kepulauan Riau",
    "Jambi", "Bengkulu", "Sumatera Selatan", "Kepulauan Bangka Belitung",
    "Lampung", "Banten", "Daerah Khusus Ibukota Jakarta", "Jawa Barat",
    "Jawa Tengah", "Daerah Istimewa Yogyakarta", "Jawa Timur",
    "Kalimantan Barat", "Kalimantan Tengah",
}
WITA = {
    "Bali", "Nusa Tenggara Barat", "Nusa Tenggara Timur",
    "Kalimantan Selatan", "Kalimantan Timur", "Kalimantan Utara",
    "Sulawesi Utara", "Sulawesi Tengah", "Sulawesi Selatan",
    "Sulawesi Tenggara", "Sulawesi Barat", "Gorontalo",
}
WIT = {
    "Maluku", "Maluku Utara", "Papua", "Papua Barat",
    "Papua Barat Daya", "Papua Pegunungan", "Papua Selatan", "Papua Tengah",
}

def zona_waktu(provinsi):
    # Ditetapkan berdasarkan pembagian administratif provinsi (geografi resmi
    # Indonesia), BUKAN dihitung dari nilai lat/lng. UTC offset disertakan
    # eksplisit agar tidak ambigu saat dipakai untuk hitungan waktu/falak.
    if provinsi in WIB:
        return {"zona": "WIB", "utc_offset": 7}
    if provinsi in WITA:
        return {"zona": "WITA", "utc_offset": 8}
    if provinsi in WIT:
        return {"zona": "WIT", "utc_offset": 9}
    if provinsi == "Arab Saudi":
        return {"zona": "Waktu Arab Saudi (AST)", "utc_offset": 3}
    return None

def clean_dms(s):
    if s is None:
        return None
    return str(s).strip()

def bounding_box_ok(lat, lng, provinsi):
    # Longgar: Indonesia + toleransi Arab Saudi untuk titik referensi Ka'bah
    if provinsi == "Arab Saudi":
        return True
    return -11.5 <= lat <= 6.5 and 94.5 <= lng <= 141.5

wb = load_workbook(SRC, read_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data_rows = rows[1:]

referensi = []
kecamatan = []
warnings = []
seen_keys = {}

for idx, row in enumerate(data_rows, start=2):  # baris excel mulai dari 2
    no, kec, kab, prov, lat, lng, lat_dms, lng_dms, elev = row

    if kec is None or prov is None:
        warnings.append(f"Baris {idx}: data kosong, dilewati")
        continue

    try:
        lat_f = float(lat)
        lng_f = float(lng)
    except (TypeError, ValueError):
        warnings.append(f"Baris {idx}: koordinat tidak valid untuk '{kec}'")
        continue

    elev_val = None
    if elev is not None:
        m = re.search(r"-?\d+(\.\d+)?", str(elev))
        if m:
            elev_val = round(float(m.group()))
        else:
            warnings.append(f"Baris {idx}: elevasi tidak terbaca untuk '{kec}' -> disimpan null")

    entry = {
        "id": None,  # diisi belakangan setelah dedupe
        "kecamatan": str(kec).strip(),
        "kabupaten": str(kab).strip() if kab else None,
        "provinsi": str(prov).strip(),
        "lat": round(lat_f, 7),
        "lng": round(lng_f, 7),
        "lat_dms": clean_dms(lat_dms),
        "lng_dms": clean_dms(lng_dms),
        "elevasi_m": elev_val,
    }
    zw = zona_waktu(str(prov).strip())
    if zw:
        entry["zona_waktu"] = zw["zona"]
        entry["utc_offset"] = zw["utc_offset"]
    else:
        entry["zona_waktu"] = None
        entry["utc_offset"] = None

    if not bounding_box_ok(lat_f, lng_f, entry["provinsi"]):
        warnings.append(f"Baris {idx}: koordinat di luar bounding box wilayah Indonesia untuk '{kec}' ({lat_f}, {lng_f})")

    # Dua titik referensi tetap (No = 0) dipisah dari daftar kecamatan
    if no == 0 or entry["kecamatan"] in ("Ka'bah (Masjidil Haram, Makkah)", "Pondok Pesantren Lirboyo"):
        referensi.append(entry)
        continue

    key = (entry["kecamatan"].lower(), entry["kabupaten"].lower() if entry["kabupaten"] else "", entry["provinsi"].lower())
    if key in seen_keys:
        warnings.append(f"Baris {idx}: kemungkinan duplikat '{kec}' / '{kab}' / '{prov}' (sama dengan baris {seen_keys[key]})")
    else:
        seen_keys[key] = idx

    kecamatan.append(entry)

# assign sequential ids
for i, e in enumerate(referensi, start=1):
    e["id"] = f"ref_{i:02d}"
for i, e in enumerate(kecamatan, start=1):
    e["id"] = f"kec_{i:05d}"

output = {
    "meta": {
        "nama_dataset": "data_koordinat",
        "sumber": "Master_Kecamatan_Tervalidasi_v7_Final.xlsx",
        "version": "v7",
        "updated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "total_referensi": len(referensi),
        "total_kecamatan": len(kecamatan),
    },
    "referensi": referensi,
    "kecamatan": kecamatan,
}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Referensi tetap: {len(referensi)}")
print(f"Total kecamatan: {len(kecamatan)}")
print(f"Total warning validasi: {len(warnings)}")
print("--- contoh warning (maks 30) ---")
for w in warnings[:30]:
    print(w)

with open("validasi_warnings.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(warnings))
